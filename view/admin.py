# -*-codeing = utf-8 -*-
#@Time:2023 05 09
#@Author:Qu Linyi
#@File:admin.py
#@Software: PyCharm
import pandas as pd
import os

from django.shortcuts import render,redirect,HttpResponse
from app01.models import User,Title,Uploadtime
from django import forms
from app01.utils.encrypt import md5
from django.core.exceptions import ValidationError

from app01.utils.pagination import Pagination
from app01.utils.bootstrap import BootStrapModelForm

import json
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse

from openpyxl import load_workbook


class TitleModelForm(forms.ModelForm):
    title_name=forms.CharField(min_length=5,label="选题名称")
    class Meta:
        model=Title
        fields=["title_name","title_quality",
                "fit_major","work_time","work_difficulty",
                "title_source","title_introduce",
                "paper_requirement","advice"]
        # widgets={
        #     "title_introduce":forms.Textarea()
        # }
    def __init__(self,*args,**kwargs):
        super().__init__(*args,**kwargs)
        #循环找到所有的插件，添加类名
        for name, field in self.fields.items():
            field.widget.attrs={'class':'form-control','placeholder':field.label}
            if name=="title_introduce" or name=="paper_requirement":
                field.widget.attrs = {'class': 'form-control h-200', 'placeholder': field.label}


def apply_list(request):
    """"选题表申请"""
    data_dict = {}  # 空字典就是查询所有
    search_data = request.GET.get("q", "")
    if search_data:   #如果传过来的search_data不为空
        data_dict["title_name__contains"]=search_data

    title_list = Title.objects.filter(**data_dict,state="待审核");
    page_object = Pagination(request, title_list)
    form = TitleModelForm()
    context = {
        "title_list": page_object.page_queryset,
        "search_data": search_data,
        "page_string": page_object.html(),
        "form": form
    }

    return render(request,"admin_applylist.html",context)

def approve_title(request,nid):
    Title.objects.filter(id=nid).update(state="通过")
    return redirect("/admin/applylist/")

def negative_title(request,nid):
    Title.objects.filter(id=nid).update(state="拒绝")
    return redirect("/admin/applylist/")

class UserModelForm(BootStrapModelForm):
    confirm_password=forms.CharField(
        label="确认密码",
        widget=forms.PasswordInput(render_value=True)
    )
    class Meta:
        model=User
        fields=['name','identity','age','password','confirm_password']
        widgets={
            'password':forms.PasswordInput(render_value=True)
            #'password': forms.PasswordInput(render_value=True)  确认密码错误后不会清空该输入框
        }
    def clean_password(self):
        pwd=self.cleaned_data.get('password')
        return md5(pwd)
    def clean_confirm_password(self):
        confirm_password=md5(self.cleaned_data.get('confirm_password'))
        password=self.cleaned_data.get('password')
        if confirm_password!=password:
            raise ValidationError("密码不一致")  #抛出异常
        #此字段返回的值可以保存到数据库，可在此处对数据进行额外处理
        return confirm_password                 #form.cleaned_data中的值  {'name':‘张三’,'confirm_password':'xxx'}
    #钩子返回加密过的密码


""""学生用户管理"""
def user_manage(request):
    if request.session['info']['id'] != 11:
        return redirect('/admin/table/')

    pagetitle="学生用户批量上传"
    data_dict = {}  # 空字典就是查询所有
    search_data = request.GET.get("q", "")
    if search_data:  # 如果传过来的search_data不为空
        data_dict["name__contains"] = search_data

    user=User.objects.filter(**data_dict,identity=2 ).exclude(id=request.session["info"]["id"])

    page_object = Pagination(request, user)
    form=UserModelForm()
    context = {
        "all_user": page_object.page_queryset,
        "search_data": search_data,
        "page_string": page_object.html(),
        "form":form,
        "pagetitle":pagetitle

    }

    return render(request,"admin_usermanage.html",context)

def stu_excel(request):
    stu_account=list(User.objects.filter(identity=2).values("id","name").all())
    df = pd.DataFrame(stu_account)
    path=os.path.join(os.path.expanduser("~"), "Desktop")
    file_path= path.replace('\\', '\\\\')
    file_path=file_path+"\\\\学生账号.xlsx"
    df.to_excel(file_path, index=False)
    return redirect('/admin/usermanage/')


""""教师用户管理"""
def teacher_user_manage(request):
    if request.session['info']['id'] != 11:
        return redirect('/admin/table/')


    pagetitle = "教师用户批量上传"
    data_dict = {}  # 空字典就是查询所有
    search_data = request.GET.get("q", "")
    if search_data:  # 如果传过来的search_data不为空
        data_dict["name__contains"] = search_data

    user = User.objects.filter(**data_dict, identity=1)

    page_object = Pagination(request, user)
    form = UserModelForm()
    context = {
        "all_user": page_object.page_queryset,
        "search_data": search_data,
        "page_string": page_object.html(),
        "form": form,
        "pagetitle": pagetitle

    }

    return render(request, "admin_teacherusermanage.html", context)

def teach_excel(request):
    stu_account=list(User.objects.filter(identity=1).values("id","name").all())
    df = pd.DataFrame(stu_account)
    path=os.path.join(os.path.expanduser("~"), "Desktop")
    file_path= path.replace('\\', '\\\\')
    file_path=file_path+"\\\\教师账号.xlsx"
    df.to_excel(file_path, index=False)
    return redirect('/admin/teacherusermanage/')


""""管理员用户管理"""
def admin_user_manage(request):
    if request.session['info']['id'] != 11:
        return redirect('/admin/table/')


    pagetitle = "admin用户批量上传"
    data_dict = {}  # 空字典就是查询所有
    search_data = request.GET.get("q", "")
    if search_data:  # 如果传过来的search_data不为空
        data_dict["name__contains"] = search_data

    user = User.objects.filter(**data_dict, identity=3).exclude(id=request.session["info"]["id"])

    page_object = Pagination(request, user)
    form = UserModelForm()
    context = {
        "all_user": page_object.page_queryset,
        "search_data": search_data,
        "page_string": page_object.html(),
        "form": form,
        "pagetitle": pagetitle

    }
    return render(request, "admin_adminusermanage.html", context)




def user_add(request):
    if request.method == 'GET':
        form=UserModelForm()
        return render(request,'admin_change.html',{"form":form,"pagetitle":"添加用户"})
    form=UserModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        return redirect('/admin/usermanage/')
    return render(request,'admin_change.html',{"form":form,"pagetitle":"添加用户"})




class AdminEditModelForm(BootStrapModelForm):
    class Meta:
        model = User
        fields = ['name', 'identity', 'age']

def edit_user(request,nid):
    row_object=User.objects.filter(id=nid).first()
    #print(row_object.get_identity_display())
    # if not row_object:
    #     return redirect('/admin/usermanage/')

    pagetitle="编辑用户"
    if request.method == 'GET':
        form=AdminEditModelForm(instance=row_object)
        return render(request,'admin_change.html',{'form':form,'pagetitle':pagetitle})
    form=AdminEditModelForm(data=request.POST,instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('/admin/usermanage/')
    return render(request,'admin_change.html',{'form':form,'pagetitle':pagetitle})

def delete_user(request,nid):
    User.objects.filter(id=nid).delete()
    return redirect('/admin/usermanage/')



def title_manage(request):
    data_dict = {}  # 空字典就是查询所有
    search_data = request.GET.get("q", "")
    if search_data:  # 如果传过来的search_data不为空
        data_dict["title_name__contains"] = search_data

    title = Title.objects.filter(**data_dict,state="通过")

    page_object = Pagination(request, title)
    form = TitleModelForm()
    context = {
        "title_list": page_object.page_queryset,
        "search_data": search_data,
        "page_string": page_object.html(),
        "form": form

    }

    return render(request, "admin_titlemanage.html", context)


"""ajax模态框添加新用户"""
@csrf_exempt
def ajax_user_add(request):
    form=UserModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        return JsonResponse({"status":True,'error':form.errors})
    return JsonResponse({"status":False,'error':form.errors})

"""ajax删除用户"""
def user_delete(request):
    uid=request.GET.get("uid")
    """校验数据是否存在"""
    exists = User.objects.filter(id=uid).exists()
    if not exists:
        return JsonResponse({'status': False, 'error': "删除失败，数据不存在。"})
    User.objects.filter(id=uid).delete()
    return JsonResponse({'status': True})
"""根据ID显示用户详细信息"""
def user_detail(request):
    """"方式 1
    uid=request.GET.get("uid")
    row_object = User.objects.filter(id=uid).first()
    if not row_object:
        return JsonResponse({'status': False, 'error': "用户不存在。"})

    #获得的对象无法直接 JSON 序列化返回
    result={
        'status': True,
        'data':{
            "name": row_object.name,
            "age": row_object.age,
            "identity": row_object.identity,
            "password": row_object.password,
        }
    }
    return JsonResponse(result)
    """
    #方式 2
    uid = request.GET.get("uid")
    row_dict = User.objects.filter(id=uid).values().first()     #此时结果直接为字典
    print(row_dict)
    if not row_dict:
        return JsonResponse({'status': False, 'error': "用户不存在。"})

    # 获得的对象无法直接 JSON 序列化返回
    result = {
        'status': True,
        'data': row_dict
    }
    return JsonResponse(result)

""""ajax编辑用户保存"""
@csrf_exempt
def user_edit(request):
    uid=request.GET.get("uid")
    row_object=User.objects.filter(id=uid).first()
    if not row_object:
        return JsonResponse({'status': False, 'tips': "编辑用户失败，数据不存在,刷新重试。"})
    #表单验证
    form=UserModelForm(data=request.POST,instance=row_object)
    if form.is_valid():
        print(form.cleaned_data)
        form.save()
        return JsonResponse({'status': True})
    return JsonResponse({'status': False ,'error': form.errors})

def password_reset(request):
    uid=request.GET.get("uid")
    resetPassword=md5("123456")
    row_object=User.objects.filter(id=uid).first()
    if not row_object:
        return JsonResponse({'status': False, 'error': "重置失败，无此用户。"})
    User.objects.filter(id=uid).update(password=resetPassword)
    return JsonResponse({'status': True})

def apply_title_detail(request):
    uid=request.GET.get("uid")
    row_dict = Title.objects.filter(id=uid).values().first()  # 此时结果直接为字典
    if not row_dict:
        return JsonResponse({'status': False, 'error': "选题不存在。"})

    # 获得的对象无法直接 JSON 序列化返回
    print(row_dict)
    result = {
        'status': True,
        'data': row_dict
    }
    return JsonResponse(result)

@csrf_exempt
def apply_title_edit(request):
    uid = request.GET.get("uid")
    row_object = Title.objects.filter(id=uid).first()
    if not row_object:
        return JsonResponse({'status': False, 'tips': "编辑选题失败，数据不存在,刷新重试。"})
    # 表单验证
    form = TitleModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return JsonResponse({'status': True})
    return JsonResponse({'status': False, 'error': form.errors})



"""管理已通过的选题"""



def title_detail(request):
    uid=request.GET.get("uid")
    row_dict = Title.objects.filter(id=uid).values().first()  # 此时结果直接为字典
    if not row_dict:
        return JsonResponse({'status': False, 'error': "选题不存在。"})

    # 获得的对象无法直接 JSON 序列化返回
    print(row_dict)
    result = {
        'status': True,
        'data': row_dict
    }
    return JsonResponse(result)

@csrf_exempt
def title_edit(request):
    uid = request.GET.get("uid")
    row_object = Title.objects.filter(id=uid).first()
    if not row_object:
        return JsonResponse({'status': False, 'tips': "编辑选题失败，数据不存在,刷新重试。"})
    # 表单验证
    form = TitleModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return JsonResponse({'status': True})
    return JsonResponse({'status': False, 'error': form.errors})

def title_delete(request):
        uid = request.GET.get("uid")
        print(uid)
        """校验数据是否存在"""
        exists = Title.objects.filter(id=uid).exists()
        if not exists:
            return JsonResponse({'status': False, 'error': "删除失败，数据不存在。"})
        Title.objects.filter(id=uid).delete()
        return JsonResponse({'status': True})


def table(request):

    return render(request,"admin_showtable.html")
def get_table_data(request):
    idlist=User.objects.filter(identity=1).values("name","id") # [{'name': '王老师2', 'id': 1}, {'name': '李老师', 'id': 9}]
    serielist=[] #封装
    state_legend = ["待审核", "通过", "拒绝"] #每个x轴的值显示几个状态的值
    x_name = [] #教师姓名列表   x轴
    for i in list(idlist):
        x_name.append(i['name'])
    wait_data=[]
    aprove_data=[]
    defuse_data=[]

    for i in list(idlist):
        wait_data.append(Title.objects.filter(teacher_id=i['id'],state="待审核").count())
        aprove_data.append(Title.objects.filter(teacher_id=i['id'],state="通过").count())
        defuse_data.append(Title.objects.filter(teacher_id=i['id'],state="拒绝").count())

    for i in state_legend:
        dic={}
        dic["name"]=i
        dic["type"]="bar"
        serielist.append(dic)
    serielist[0]["data"]=wait_data
    serielist[1]["data"]=aprove_data
    serielist[2]["data"]=defuse_data


    # # 假设下面是数据库获取到的信息
    # legend = ["甲", "乙"]
    # series_list = [
    #     {
    #         "name": '待审核',
    #         "type": 'bar',
    #         "data": [5, 20, 36, 10, 10, 20]
    #     },
    #     {
    #         "name": '通过',
    #         "type": 'bar',
    #         "data": [15, 40, 66, 40, 50, 10]
    #     }
    # ]
    # x_axis = ['王老师', '2月', '3月', '4月', '5月', '6月']
    result = {
        'status': True,
        'data': {
            'legend': state_legend,
            'series_list': serielist,
            "x_axis": x_name
        }
    }

    return JsonResponse(result)

def get_pie_data(request):
    waitapprovenum=Title.objects.filter(state="待审核").count()
    approvenum=Title.objects.filter(state="通过").count()
    defusenum=Title.objects.filter(state="拒绝").count()
    db_data=[
        {"value": waitapprovenum,"name": '待审核'},
        {"value": approvenum,"name": '已通过'},
        {"value": defusenum,"name": '拒绝'}
    ]
    result = {
        'status': True,
        'data': db_data
    }
    return JsonResponse(result)

""""基于Excel批量添加学生用户"""
def multi_user_upload(request):
    first_password=md5("123456")
    file_object=request.FILES.get('exc')
    print(type(file_object))

    # 2.对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]#第一行
    for row in sheet.iter_rows(min_row=2):
        # 第一列的值
        text = row[0].value
        age=row[1].value
        exists = User.objects.filter(name=text).exists()
        # 不存在才添加
        if not exists:
            User.objects.create(name=text,password=first_password,identity=2,age=age)


    return redirect('/admin/usermanage/')


def multi_teacheruser_upload(request):
    first_password = md5("123456")
    file_object = request.FILES.get('exc')
    print(type(file_object))

    # 2.对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]  # 第一行
    for row in sheet.iter_rows(min_row=2):
        # 第一列的值
        text = row[0].value
        age = row[1].value
        exists = User.objects.filter(name=text).exists()
        # 不存在才添加
        if not exists:
            User.objects.create(name=text, password=first_password, identity=1, age=age)

    return redirect('/admin/teacherusermanage/')


# class UserModelForm(BootStrapModelForm):
#     confirm_password=forms.CharField(
#         label="确认密码",
#         widget=forms.PasswordInput(render_value=True)
#     )
#     class Meta:
#         model=User
#         fields=['name','identity','age','password','confirm_password']
#         widgets={
#             'password':forms.PasswordInput(render_value=True)
#             #'password': forms.PasswordInput(render_value=True)  确认密码错误后不会清空该输入框
#         }
#     def clean_password(self):
#         pwd=self.cleaned_data.get('password')
#         return md5(pwd)
#     def clean_confirm_password(self):
#         confirm_password=md5(self.cleaned_data.get('confirm_password'))
#         password=self.cleaned_data.get('password')
#         if confirm_password!=password:
#             raise ValidationError("密码不一致")  #抛出异常
#         #此字段返回的值可以保存到数据库，可在此处对数据进行额外处理
#         return confirm_password                 #form.cleaned_data中的值  {'name':‘张三’,'confirm_password':'xxx'}
#     #钩子返回加密过的密码

from django.core.validators import RegexValidator
class UserMyInfoModelForm(BootStrapModelForm):
    tele = forms.CharField(
        label='手机号',
        validators=[RegexValidator(r'^1[3-9]\d{9}$', '手机格式错误')],
    )
    class Meta:
        model=User
        fields=["name","age","email","tele"]

def my_information(request,nid):
    row_object=User.objects.filter(id=nid).first()
    title="我的信息"
    if request.method == "GET":
        form=UserMyInfoModelForm(instance=row_object)
        return render(request,'admin_change.html',{"form":form,"pagetitle":title})
    form=UserMyInfoModelForm(data=request.POST,instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('/admin/table/')
    return render(request,'admin_change.html',{'form':form,'pagetitle':title})

class UserChangePasswordModelForm(BootStrapModelForm):
    password_now=forms.CharField(
        label="当前密码",
        widget=forms.PasswordInput(render_value=False)
    )
    confirm_password = forms.CharField(
            label="确认密码",
            widget=forms.PasswordInput(render_value=True)
        )
    class Meta:
        model=User
        fields=['password_now','password','confirm_password']
        widgets = {
                    'password':forms.PasswordInput(render_value=True)
                    #'password': forms.PasswordInput(render_value=True)  确认密码错误后不会清空该输入框
                }


    def clean_password(self):
        pwd=self.cleaned_data.get('password')

        return md5(pwd)
    def clean_confirm_password(self):
        confirm_password=md5(self.cleaned_data.get('confirm_password'))
        password=self.cleaned_data.get('password')
        if confirm_password!=password:
            raise ValidationError("密码不一致")  #抛出异常
        #此字段返回的值可以保存到数据库，可在此处对数据进行额外处理
        return confirm_password                 #form.cleaned_data中的值  {'name':‘张三’,'confirm_password':'xxx'}
    #钩子返回加密过的密码
    def clean_password_now(self):
        pwd = self.cleaned_data.get('password_now')
        md5_pwd = md5(pwd)
        # 去数据库校验当前密码和新输入的密码是否一致
        exists = User.objects.filter(id=self.instance.pk, password=md5_pwd).exists()
        if not exists:
            raise ValidationError('密码错误')

def change_password(request,nid):
    row_object = User.objects.filter(id=nid).first()
    title = "修改密码"
    if request.method == "GET":
        form = UserChangePasswordModelForm()
        return render(request, 'admin_change.html', {"form": form, "pagetitle": title})
    form = UserChangePasswordModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('/admin/table/')
    return render(request, 'admin_change.html', {'form': form, 'pagetitle': title})



class SetUploadTime(BootStrapModelForm):
    class Meta:
        model=Uploadtime
        fields=["start_time","end_time"]
    def clean_end_time(self):
        start_date = self.cleaned_data.get('start_time')
        end_date = self.cleaned_data.get('end_time')

        if start_date and end_date:
            if start_date > end_date:
                raise ValidationError("开始日期不能晚于结束日期")
        return end_date

def uploadtime_manage_page(request):
    uploadset_list=Uploadtime.objects.all()
    form=SetUploadTime()
    return render(request,'admin_uploadtime_manage.html',{"title_list":uploadset_list,"form":form})

def set_uploadtime_detail(request):
    uid=request.GET.get("uid")
    row_object=Uploadtime.objects.filter(id=uid).values().first()
    if not row_object:
        return JsonResponse({'status': False, 'error': "数据不存在,刷新重试。"})
    return JsonResponse({'status': True,"data":row_object})

@csrf_exempt
def set_uploadtime_save(request):
    uid=request.GET.get("uid")
    row_object = Uploadtime.objects.filter(id=uid).first()
    if not row_object:
        return JsonResponse({'status': False, 'tips': "数据不存在,刷新重试。"})
    form=SetUploadTime(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return JsonResponse({'status': True})
    return JsonResponse({'status': False, 'error': form.errors})
